import time

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import load_workbook

print("Iniciando nosso robô...\n")

domains = []
file = open('results.csv', 'w')

# Reading by excel
wb = load_workbook('domains.xlsx')
sheet = wb.active

for value in sheet.iter_rows(min_row=1, max_row=7, values_only=True):
    domains.append(value)

options = Options()
options.add_argument("start-maximized")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get("https://registro.br/")

# items = driver.find_elements(By.XPATH, '//li')
# for item in items:
#     print(item.text)


# domains = ["rhsm.com.br", "udemy.com", "uol.com.br", "globo.com.br", "sixround123.com"]

for domain in domains:
    search = driver.find_element(By.ID, "is-avail-field")
    search.clear()
    search.send_keys(domain)
    search.send_keys(Keys.ENTER)
    time.sleep(1)

    result = driver.find_elements(By.TAG_NAME, "strong")

    text = f"{domain[0]}, {result[4].text}\n"
    print(text)
    file.write(text)
    time.sleep(1)

file.close()
# for i, v in enumerate(result):
#     print(i, v.text)
driver.close()
